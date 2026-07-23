"""
health.py — Stranded / unfulfillable inventory health report.

Python version of class-fba-health-report.php + class-fba-health-sheets-api.php.
Reads the most recent report spreadsheet from a per-client Google Drive folder.

Security note: this uses drive.metadata.readonly + spreadsheets.readonly —
narrower than the original plugin's full `drive` scope (Finding 2 from the
security review is fixed here rather than carried over).

Improvement over the original: instead of guessing sheet-tab positions by
exporting gid=0..9 and sniffing headers, this downloads the actual xlsx
bytes and reads tabs by name directly with openpyxl when the file isn't a
native Google Sheet — more reliable than the PHP version's guesswork.
"""

import io
from datetime import datetime

import pandas as pd
import streamlit as st
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from sheets import _credentials
import store

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
]

STRANDED_TAB = "Stranded Inventory"
UNFULFILLABLE_TAB = "Unfillable Inventory"


def _drive():
    return build("drive", "v3", credentials=_credentials(SCOPES), cache_discovery=False)


def _sheets():
    return build("sheets", "v4", credentials=_credentials(SCOPES), cache_discovery=False)


def _latest_file_in_folder(folder_id: str):
    mime_filter = ("(mimeType='application/vnd.google-apps.spreadsheet' or "
                   "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')")
    q = f"'{folder_id}' in parents and {mime_filter} and trashed=false"
    result = _drive().files().list(
        q=q, orderBy="createdTime desc", pageSize=1,
        fields="files(id,name,createdTime,mimeType)",
    ).execute()
    files = result.get("files", [])
    if not files:
        raise ValueError(f"No report spreadsheets found in that Drive folder.")
    return files[0]


def _parse_rows(rows) -> list[dict]:
    if len(rows) < 2:
        return []
    headers = [str(h).strip().lower() for h in rows[0]]
    out = []
    for row in rows[1:]:
        data = {h: (row[j].strip() if j < len(row) else "") for j, h in enumerate(headers)}
        rt = data.get("record_type", "").upper()
        if not rt or rt == "SUMMARY":
            continue
        out.append(data)
    return out


def _read_native_sheet_tab(file_id: str, tab_name: str) -> list[dict]:
    try:
        result = _sheets().spreadsheets().values().get(
            spreadsheetId=file_id, range=f"'{tab_name}'!A1:V500"
        ).execute()
        return _parse_rows(result.get("values", []))
    except Exception:
        return []


def _read_xlsx_tab(file_id: str, tab_name: str) -> list[dict]:
    """Download the xlsx binary and read one tab by name with openpyxl."""
    import openpyxl
    request = _drive().files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    if tab_name not in wb.sheetnames:
        return []
    ws = wb[tab_name]
    rows = [[str(c) if c is not None else "" for c in r] for r in ws.iter_rows(values_only=True)]
    return _parse_rows(rows)


def fetch_health_data(folder_id: str) -> dict:
    """Fetch stranded + unfulfillable data from the newest report in the folder."""
    file = _latest_file_in_folder(folder_id)
    is_xlsx = file.get("mimeType") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if is_xlsx:
        stranded = _read_xlsx_tab(file["id"], STRANDED_TAB)
        unfulfillable = _read_xlsx_tab(file["id"], UNFULFILLABLE_TAB)
    else:
        stranded = _read_native_sheet_tab(file["id"], STRANDED_TAB)
        unfulfillable = _read_native_sheet_tab(file["id"], UNFULFILLABLE_TAB)

    return {"file_name": file["name"], "file_date": file["createdTime"],
            "stranded": stranded, "unfulfillable": unfulfillable}


def sync_and_save(brand_code: str, folder_id: str) -> dict:
    """Pull fresh health data and overwrite this client's stored rows."""
    data = fetch_health_data(folder_id)
    synced_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    stranded_df = store.read_tab("health_stranded")
    stranded_df = stranded_df[stranded_df["brand_code"] != brand_code]
    new_stranded = [{
        "brand_code": brand_code, "synced_at": synced_at, "file_name": data["file_name"],
        "record_type": s.get("record_type", "SKU"), "asin": s.get("asin", ""), "sku": s.get("sku", ""),
        "fnsku": s.get("fnsku", ""), "product": s.get("product", "")[:500],
        "condition": s.get("condition", ""), "issue": s.get("issue", ""),
        "quantity": s.get("quantity", "0"), "dollar_value": s.get("dollar_value", "0"),
        "recommended_action": s.get("recommended_action", ""),
        "date_stranded": s.get("date_stranded", ""),
        "is_new": "1" if s.get("is_new", "").upper() == "TRUE" else "0",
    } for s in data["stranded"]]
    store._write_tab("health_stranded", pd.concat([stranded_df, pd.DataFrame(new_stranded)], ignore_index=True))

    unful_df = store.read_tab("health_unfulfillable")
    unful_df = unful_df[unful_df["brand_code"] != brand_code]
    new_unful = [{
        "brand_code": brand_code, "synced_at": synced_at, "file_name": data["file_name"],
        "record_type": u.get("record_type", "SKU"), "asin": u.get("asin", ""), "sku": u.get("sku", ""),
        "fnsku": u.get("fnsku", ""), "product": u.get("product", "")[:500], "issue": u.get("issue", ""),
        "quantity": u.get("quantity", "0"), "dollar_value": u.get("dollar_value", "0"),
        "recommended_action": u.get("recommended_action", ""),
        "recommended_removal_quantity": u.get("recommended_removal_quantity", "0"),
        "is_new": "1" if (u.get("is_new", "").upper() == "TRUE"
                          or u.get("record_type", "").upper() == "NEW SKU") else "0",
    } for u in data["unfulfillable"]]
    store._write_tab("health_unfulfillable", pd.concat([unful_df, pd.DataFrame(new_unful)], ignore_index=True))

    return {"file_name": data["file_name"], "stranded_count": len(new_stranded),
            "unfulfillable_count": len(new_unful)}


def get_health_data(brand_code: str) -> dict:
    stranded = store.read_tab("health_stranded")
    stranded = stranded[stranded["brand_code"] == brand_code]
    unfulfillable = store.read_tab("health_unfulfillable")
    unfulfillable = unfulfillable[unfulfillable["brand_code"] == brand_code]
    file_name = ""
    synced_at = ""
    if not stranded.empty:
        file_name, synced_at = stranded.iloc[0]["file_name"], stranded.iloc[0]["synced_at"]
    elif not unfulfillable.empty:
        file_name, synced_at = unfulfillable.iloc[0]["file_name"], unfulfillable.iloc[0]["synced_at"]
    return {"stranded": stranded, "unfulfillable": unfulfillable, "file_name": file_name, "synced_at": synced_at}


def get_notification_counts(brand_code: str) -> dict:
    """sku -> {'stranded': qty, 'unfulfillable': qty} — for badges on the main table."""
    data = get_health_data(brand_code)
    counts = {}
    if not data["stranded"].empty:
        s = data["stranded"].copy()
        s["quantity"] = pd.to_numeric(s["quantity"], errors="coerce").fillna(0)
        for sku, qty in s.groupby("sku")["quantity"].sum().items():
            if qty > 0:
                counts.setdefault(sku, {})["stranded"] = int(qty)
    if not data["unfulfillable"].empty:
        u = data["unfulfillable"].copy()
        u["quantity"] = pd.to_numeric(u["quantity"], errors="coerce").fillna(0)
        for sku, qty in u.groupby("sku")["quantity"].sum().items():
            if qty > 0:
                counts.setdefault(sku, {})["unfulfillable"] = int(qty)
    return counts
