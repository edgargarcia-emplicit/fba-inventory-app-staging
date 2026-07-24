"""
xlsx_export_parser.py — Parses an Amazon listings export (.xlsx) into
per-ASIN content maps (title, bullets 1-10, description, keywords).

Handles BOTH the new Amazon bulk-listing flat-file format (columns like
`item_name__1__value`, `bullet_point__1__value` .. `__10__value`, double
underscores) and the older, simpler column-name format the original
WordPress plugin was built against (`itemName`, `bullet_point.0.value`,
dot notation, 5 bullets only) — since existing files in that older shape
may still show up.
"""

import re

from openpyxl import load_workbook

ASIN_RE = re.compile(r"^B[A-Z0-9]{9}$")


def _normalize_header(h) -> str:
    return re.sub(r"\s+", " ", str(h or "")).strip().lower()


def _map_header_to_field(header: str) -> str | None:
    """Map a column header (already lowercased) to an internal field name."""
    header = header.strip()

    direct_map = {
        "asin": "asin",
        "sku": "sku",
        "itemname": "title",
        "item_name": "title",
        "title": "title",
        "product_name": "title",
        "item_name__1__value": "title",
        "brand": "brand",
        "brand.0.value": "brand",
        "brand__1__value": "brand",
        "product_description.0.value": "description",
        "product_description": "description",
        "product_description__1__value": "description",
    }
    if header in direct_map:
        return direct_map[header]

    # New format: bullet_point__1__value .. __10__value
    m = re.match(r"^bullet_point__(\d+)__value$", header)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 10:
            return f"bullet_{n}"

    # Old format: bullet_point.0.value (zero-indexed) or bulletN
    m = re.match(r"^bullet_point\.(\d+)\.value$", header)
    if m:
        return f"bullet_{int(m.group(1)) + 1}"
    m = re.match(r"^bullet(\d)$", header)
    if m:
        return f"bullet_{m.group(1)}"
    m = re.match(r"^bullet_(\d)$", header)
    if m:
        return f"bullet_{m.group(1)}"

    # Keywords — new format generic_keyword__N__value, old dot notation
    if re.match(r"^generic_keyword__\d+__value$", header):
        return "keyword"
    if re.match(r"^generic_keyword\.\d+\.value$", header):
        return "keyword"
    if re.match(r"^subject_keyword\.\d+\.value$", header):
        return "keyword"

    return None


NUM_BULLETS = 10
BASE_FIELDS = ["asin", "sku", "title", "brand", "description"] + [f"bullet_{i}" for i in range(1, NUM_BULLETS + 1)]


def parse(file_path: str) -> dict:
    """
    Parse an Amazon listings .xlsx export. Scans all sheets.

    Returns {ASIN: {asin, skus: [...], title, brand, bullet_1..10,
                    description, keywords}}
    Keeps the row with the most populated fields when the same ASIN
    appears on multiple rows (e.g. a parent row + several offer rows).
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)

    by_asin: dict[str, dict] = {}
    skus_for_asin: dict[str, list] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_map = {}  # col_index -> field
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        for col_idx, header in enumerate(header_row):
            field = _map_header_to_field(_normalize_header(header))
            if field:
                header_map[col_idx] = field

        if not header_map:
            continue

        for row in rows_iter:
            parsed = {f: "" for f in BASE_FIELDS}
            keywords = []
            for col_idx, field in header_map.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or val == "":
                    continue
                val = str(val).strip()
                if field == "keyword":
                    keywords.append(val)
                else:
                    parsed[field] = val

            asin = parsed["asin"].strip().upper()
            if not asin or not ASIN_RE.match(asin):
                continue

            if parsed.get("sku"):
                skus_for_asin.setdefault(asin, [])
                if parsed["sku"] not in skus_for_asin[asin]:
                    skus_for_asin[asin].append(parsed["sku"])

            score = sum(1 for f in ("title", "bullet_1", "bullet_2", "bullet_3", "description") if parsed.get(f))
            score += len(keywords)

            if asin not in by_asin or score > by_asin[asin].get("_score", -1):
                parsed["_score"] = score
                parsed["_keywords"] = keywords
                by_asin[asin] = parsed

    result = {}
    for asin, row in by_asin.items():
        row = dict(row)
        row.pop("_score", None)
        keywords = row.pop("_keywords", [])
        row["keywords"] = "; ".join(k for k in keywords if k)
        row["skus"] = skus_for_asin.get(asin, ([row["sku"]] if row.get("sku") else []))
        result[asin] = row

    if not result:
        raise ValueError(
            "No valid rows found in the Excel file. Make sure it has an 'asin' column "
            "and a title column (item_name__1__value or itemName)."
        )
    return result
